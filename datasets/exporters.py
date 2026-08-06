import csv
import io
import logging
import re
import zipfile
from collections import Counter
from datetime import datetime
from pathlib import Path

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import pandas as pd
from arabic_reshaper import reshape
from bidi.algorithm import get_display
from django.conf import settings
from django.utils import timezone
from matplotlib import font_manager
from matplotlib.ticker import FuncFormatter
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    Image, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)

logger = logging.getLogger(__name__)

PREVIEW_ROWS = 100
TOP_RESULTS = 10
MAX_CORRELATION_COLUMNS = 20
MAX_REPORT_CELLS = 2_000_000
ARABIC_FONT_NAME = 'AnalytixArabic'
ARABIC_FONT_FILENAMES = (
    'Cairo-Regular.ttf',
    'Tajawal-Regular.ttf',
    'IBMPlexSansArabic-Regular.ttf',
    'NotoSansArabic-Regular.ttf',
    'DejaVuSans.ttf',
)
BRAND_BLUE = '1D4ED8'
BRAND_LIGHT = 'DBEAFE'
CHART_COLORS = ['#2563EB', '#14B8A6', '#0EA5E9', '#F59E0B', '#EF4444', '#8B5CF6']


class ExportSourceError(Exception):
    pass


class ArabicFontNotFoundError(ExportSourceError):
    pass


def load_workbook_data(dataset):
    """Read the source exactly once per export request."""
    try:
        dataset.file.open('rb')
        with pd.ExcelFile(dataset.file) as excel:
            frames = {name: excel.parse(name) for name in excel.sheet_names}
        if sum(frame.size for frame in frames.values()) > MAX_REPORT_CELLS:
            raise ExportSourceError(
                'حجم البيانات أكبر من الحد المسموح لإنشاء التقرير المتزامن.'
            )
        return frames
    except ExportSourceError:
        raise
    except Exception as exc:
        logger.warning('Dataset source could not be read for export', exc_info=True)
        raise ExportSourceError(
            'تعذر الوصول إلى ملف البيانات. يرجى المحاولة لاحقًا أو إعادة رفع الملف.'
        ) from exc
    finally:
        try:
            dataset.file.close()
        except Exception:
            pass


def _dashboard_widgets(dataset):
    from dashboards.models import DashboardWidget
    return list(DashboardWidget.objects.filter(
        dashboard__owner=dataset.user,
        dashboard__layout_settings__dataset_id=dataset.pk,
        is_visible=True,
    ).select_related('dashboard', 'analysis_result'))


def _existing_results(dataset):
    from analysis.models import AnalysisResult
    results = AnalysisResult.objects.filter(
        analysis_job__dataset=dataset,
        analysis_job__owner=dataset.user,
    ).select_related('analysis_job')
    return [{
        'name': result.analysis_job.name,
        'summary': result.summary,
        'statistics': result.statistics,
        'insights': result.insights,
    } for result in results]


def _number(value, digits=2):
    if pd.isna(value):
        return None
    return round(float(value), digits)


def _severity(missing_rate=0, duplicate_rate=0, outlier_rate=0):
    score = max(missing_rate, duplicate_rate, outlier_rate)
    return 'مرتفع' if score >= 20 else ('متوسط' if score >= 5 else 'منخفض')


def _prepare_report_frame(frame):
    """Return a display-only copy with useful, readable column names."""
    display = frame.copy()
    drop_columns = []
    renamed = {}
    unnamed_index = 0
    for column in display.columns:
        if str(column).strip().lower().startswith('unnamed:'):
            if display[column].isna().all():
                drop_columns.append(column)
            else:
                unnamed_index += 1
                renamed[column] = f'عمود غير مسمى {unnamed_index}'
    if drop_columns:
        display = display.drop(columns=drop_columns)
    if renamed:
        display = display.rename(columns=renamed)
    return display


def _date_columns(frame):
    dates = []
    for column in frame.columns:
        series = frame[column]
        converted = None
        if pd.api.types.is_datetime64_any_dtype(series):
            converted = pd.to_datetime(series, errors='coerce')
        elif series.dtype == 'object' and len(series.dropna()) and len(series.dropna()) <= 10_000:
            try:
                candidate = pd.to_datetime(series, errors='coerce')
                if candidate.notna().mean() >= .8:
                    converted = candidate
            except Exception:
                pass
        if converted is not None and converted.notna().any():
            dates.append({
                'column': str(column),
                'count': int(converted.notna().sum()),
                'earliest': converted.min().isoformat(),
                'latest': converted.max().isoformat(),
            })
    return dates


def _sheet_report(name, frame):
    rows, columns = frame.shape
    cells = rows * columns
    missing_cells = int(frame.isna().sum().sum())
    completion = round((1 - missing_cells / cells) * 100, 2) if cells else 100.0
    duplicates = int(frame.duplicated().sum()) if rows else 0
    duplicate_rate = round(duplicates / rows * 100, 2) if rows else 0.0
    column_details, missing, text_analysis, numeric_stats, outliers = [], [], [], [], []

    for column in frame.columns:
        series = frame[column]
        non_null = series.dropna()
        missing_count = int(series.isna().sum())
        missing_rate = round(missing_count / rows * 100, 2) if rows else 0.0
        unique = int(non_null.nunique())
        detail = {
            'column': str(column), 'type': str(series.dtype), 'missing': missing_count,
            'missing_rate': missing_rate, 'unique': unique,
            'constant': bool(len(non_null) and unique == 1),
            'high_cardinality': bool(len(non_null) >= 20 and unique / len(non_null) >= .9),
        }
        column_details.append(detail)
        missing.append({'column': str(column), 'count': missing_count, 'rate': missing_rate})

        if pd.api.types.is_numeric_dtype(series):
            clean = pd.to_numeric(non_null, errors='coerce').dropna()
            if len(clean):
                q1, median, q3 = clean.quantile([.25, .5, .75])
                iqr = q3 - q1
                outlier_count = int(((clean < q1 - 1.5 * iqr) | (clean > q3 + 1.5 * iqr)).sum()) if iqr else 0
                outlier_rate = round(outlier_count / len(clean) * 100, 2)
                numeric_stats.append({
                    'column': str(column), 'count': int(clean.count()), 'sum': _number(clean.sum()),
                    'mean': _number(clean.mean()), 'median': _number(median), 'std': _number(clean.std()),
                    'min': _number(clean.min()), 'q1': _number(q1), 'q3': _number(q3),
                    'max': _number(clean.max()),
                })
                outliers.append({'column': str(column), 'count': outlier_count, 'rate': outlier_rate})
        elif not pd.api.types.is_datetime64_any_dtype(series):
            common = non_null.astype(str).value_counts().head(TOP_RESULTS)
            text_analysis.append({
                'column': str(column), 'unique': unique,
                'top_values': [{'value': value, 'count': int(count)} for value, count in common.items()],
            })

    numeric = frame.select_dtypes(include='number')
    correlations = []
    if 2 <= len(numeric.columns) <= MAX_CORRELATION_COLUMNS and len(frame) >= 3:
        matrix = numeric.corr()
        for left_index, left in enumerate(matrix.columns):
            for right in matrix.columns[left_index + 1:]:
                value = matrix.loc[left, right]
                if pd.notna(value) and abs(value) >= .7:
                    correlations.append({'column_1': str(left), 'column_2': str(right), 'value': round(float(value), 3)})
        correlations.sort(key=lambda item: abs(item['value']), reverse=True)
        correlations = correlations[:TOP_RESULTS]

    dates = _date_columns(frame)
    issues = []
    for item in sorted(missing, key=lambda x: x['rate'], reverse=True)[:TOP_RESULTS]:
        if item['count']:
            issues.append({'issue': f'{item["column"]}: {item["rate"]}% قيم فارغة', 'severity': _severity(missing_rate=item['rate'])})
    if duplicates:
        issues.append({'issue': f'{duplicates} صفوف مكررة ({duplicate_rate}%)', 'severity': _severity(duplicate_rate=duplicate_rate)})
    for detail in column_details:
        if detail['constant']:
            issues.append({'issue': f'{detail["column"]}: عمود ثابت', 'severity': 'منخفض'})
    for item in outliers:
        if item['count']:
            issues.append({'issue': f'{item["column"]}: {item["count"]} قيم شاذة ({item["rate"]}%)', 'severity': _severity(outlier_rate=item['rate'])})

    # Quality: 50% completeness, 25% uniqueness of rows, 15% outlier cleanliness,
    # and 10% absence of constant columns. This is deterministic and data-backed.
    outlier_total = sum(item['count'] for item in outliers)
    numeric_non_null = sum(item['count'] for item in numeric_stats)
    outlier_clean = 1 - outlier_total / numeric_non_null if numeric_non_null else 1
    constant_rate = sum(item['constant'] for item in column_details) / columns if columns else 0
    quality = round(max(0, min(100, completion * .5 + (100 - duplicate_rate) * .25 + outlier_clean * 100 * .15 + (1 - constant_rate) * 100 * .10)), 1)
    recommendations = []
    if missing_cells: recommendations.append('راجع الأعمدة الأعلى في القيم الفارغة وحدد سياسة معالجة مناسبة.')
    if duplicates: recommendations.append('تحقق من الصفوف المكررة قبل استخدامها في المؤشرات.')
    if outlier_total: recommendations.append('راجع القيم الشاذة قبل بناء النماذج أو المجاميع النهائية.')
    if any(item['constant'] for item in column_details): recommendations.append('استبعد الأعمدة الثابتة من التحليلات التي لا تستفيد منها.')
    if not recommendations: recommendations.append('لا توجد مشكلة بارزة وفق قواعد الجودة الحالية؛ استمر في التحقق الدوري.')

    return {
        'name': str(name), 'rows': rows, 'columns': columns, 'cells': cells,
        'missing_cells': missing_cells, 'completion_rate': completion,
        'duplicates': duplicates, 'duplicate_rate': duplicate_rate,
        'quality_score': quality, 'columns_info': column_details,
        'missing': sorted(missing, key=lambda x: x['rate'], reverse=True),
        'numeric_stats': numeric_stats, 'text_analysis': text_analysis,
        'date_analysis': dates, 'outliers': outliers, 'correlations': correlations,
        'issues': issues[:20], 'recommendations': recommendations[:5],
        'preview': frame.head(PREVIEW_ROWS).copy(), '_frame': frame,
    }


def build_report_data(dataset, frames):
    # Cleaning is limited to the report copy; source frames and uploaded files
    # remain unchanged.
    sheets = [
        _sheet_report(name, _prepare_report_frame(frame))
        for name, frame in frames.items()
    ]
    total_cells = sum(item['cells'] for item in sheets)
    total_missing = sum(item['missing_cells'] for item in sheets)
    completion = round((1 - total_missing / total_cells) * 100, 2) if total_cells else 100.0
    quality = round(sum(item['quality_score'] * max(item['cells'], 1) for item in sheets) /
                    sum(max(item['cells'], 1) for item in sheets), 1) if sheets else 100.0
    all_issues = [{**issue, 'sheet': sheet['name']} for sheet in sheets for issue in sheet['issues']]
    severity_order = {'مرتفع': 0, 'متوسط': 1, 'منخفض': 2}
    all_issues.sort(key=lambda x: severity_order.get(x['severity'], 3))
    recommendations = list(dict.fromkeys(rec for sheet in sheets for rec in sheet['recommendations']))[:5]
    insights = [
        f'يحتوي الملف على {len(sheets)} أوراق و{sum(x["rows"] for x in sheets):,} صفوف.',
        f'بلغ اكتمال البيانات {completion}% ودرجة الجودة العامة {quality} من 100.',
        f'إجمالي الخلايا الفارغة {total_missing:,} خلية.',
        f'تم رصد {sum(x["duplicates"] for x in sheets):,} صفوف مكررة.',
        f'تم رصد {sum(sum(o["count"] for o in x["outliers"]) for x in sheets):,} قيم شاذة وفق IQR.',
    ]
    widgets = [{
        'id': widget.pk, 'title': widget.title, 'type': widget.widget_type,
        'sheet': (widget.settings or {}).get('sheet_name', ''),
        'x_column': widget.x_column, 'y_column': widget.y_column,
        'aggregation': widget.aggregation, 'colors': (widget.settings or {}).get('colors', []),
        'settings': widget.settings or {},
    } for widget in _dashboard_widgets(dataset)]
    return {
        'dataset': {'id': dataset.pk, 'title': dataset.title, 'filename': dataset.original_filename or dataset.title, 'status': dataset.get_status_display()},
        'user': {'username': dataset.user.get_username()},
        'generated_at': timezone.localtime(), 'sheets': sheets, 'widgets': widgets,
        'existing_results': _existing_results(dataset),
        'summary': {
            'sheet_count': len(sheets), 'rows': sum(x['rows'] for x in sheets),
            'columns': sum(x['columns'] for x in sheets), 'cells': total_cells,
            'missing_cells': total_missing, 'completion_rate': completion,
            'duplicates': sum(x['duplicates'] for x in sheets), 'quality_score': quality,
            'outliers': sum(sum(o['count'] for o in x['outliers']) for x in sheets),
        },
        'executive_summary': {'insights': insights[:5], 'issues': all_issues[:5], 'recommendations': recommendations,
                              'review_sheets': [x['name'] for x in sheets if x['quality_score'] < 80][:5]},
        'warnings': [],
    }


def _safe_cell(value):
    if value is None or (not isinstance(value, (list, dict)) and pd.isna(value)): return ''
    if isinstance(value, (list, dict)): return str(value)
    if isinstance(value, pd.Timestamp): return value.to_pydatetime()
    return value.item() if hasattr(value, 'item') else value


def build_excel_report(dataset, report_data):
    """Build the multi-sheet analytical workbook, not a raw-data export."""
    report = report_data if 'summary' in report_data else build_report_data(dataset, report_data)
    workbook = Workbook()
    workbook.remove(workbook.active)
    links = []
    used_sheet_names = set()

    def unique_sheet_title(title):
        cleaned = re.sub(r'[\\/*?:\[\]]', '_', str(title)).strip() or 'ورقة'
        base = cleaned[:31]
        candidate = base
        counter = 2
        while candidate.casefold() in used_sheet_names:
            suffix = f' {counter}'
            candidate = f'{base[:31 - len(suffix)]}{suffix}'
            counter += 1
        used_sheet_names.add(candidate.casefold())
        return candidate

    def add_sheet(title, headers=None, rows=None):
        rows = rows or []
        sheet = workbook.create_sheet(unique_sheet_title(title))
        sheet.sheet_view.rightToLeft = True
        if headers:
            sheet.append(headers)
            for cell in sheet[1]:
                cell.fill = PatternFill('solid', fgColor=BRAND_BLUE); cell.font = Font(color='FFFFFF', bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, readingOrder=2)
            sheet.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{max(1, len(rows) + 1)}'
            sheet.freeze_panes = 'A2'
        for row in rows: sheet.append([_safe_cell(value) for value in row])
        for column in range(1, sheet.max_column + 1):
            letter = get_column_letter(column)
            values = [sheet.cell(row, column).value for row in range(1, sheet.max_row + 1)]
            longest = max((len(str(value)) for value in values if value is not None), default=0)
            sheet.column_dimensions[letter].width = min(45, max(12, longest + 3))
            if sheet.cell(1, column).value is not None:
                header = str(sheet.cell(1, column).value)
                for cell in list(sheet.columns)[column - 1]:
                    cell.alignment = Alignment(horizontal='right', vertical='top', wrap_text=True, readingOrder=2)
                    if isinstance(cell.value, (datetime, pd.Timestamp)):
                        cell.number_format = 'yyyy-mm-dd'
                    elif isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                        cell.number_format = '0.00"%"' if '%' in header or 'نسبة' in header else '#,##0.00'
        for row_index in range(1, sheet.max_row + 1):
            longest_text = max(
                (len(str(sheet.cell(row_index, column).value or '')) for column in range(1, sheet.max_column + 1)),
                default=0,
            )
            if longest_text > 45:
                sheet.row_dimensions[row_index].height = min(90, 15 * (1 + longest_text // 45))
        links.append((title, sheet.title))
        return sheet

    cover = add_sheet('غلاف التقرير')
    cover.append(['Analytix']); cover['A1'].font = Font(size=24, bold=True, color=BRAND_BLUE)
    for row in [['تقرير تحليل البيانات الشامل'], ['الملف', report['dataset']['filename']], ['المستخدم', report['user']['username']],
                ['Dataset ID', report['dataset']['id']], ['تاريخ الإنشاء', report['generated_at'].strftime('%Y-%m-%d %H:%M')]]: cover.append(row)
    cover.column_dimensions['A'].width = 32; cover.column_dimensions['B'].width = 45
    for row in cover.iter_rows():
        for cell in row: cell.alignment = Alignment(horizontal='right', vertical='top', wrap_text=True, readingOrder=2)
    dashboard_sheet = add_sheet('Dashboard')
    dashboard_sheet.merge_cells('A1:F1'); dashboard_sheet['A1'] = 'لوحة مؤشرات Analytix'
    dashboard_sheet['A1'].font = Font(size=22, bold=True, color='FFFFFF')
    dashboard_sheet['A1'].fill = PatternFill('solid', fgColor=BRAND_BLUE)
    dashboard_sheet['A1'].alignment = Alignment(horizontal='right', vertical='center', readingOrder=2)
    dashboard_sheet.row_dimensions[1].height = 34
    dashboard_sheet.sheet_view.showGridLines = False
    dashboard_sheet.freeze_panes = 'A3'
    executive = add_sheet('الملخص التنفيذي', ['القسم', 'التفصيل'],
        [['نتيجة', value] for value in report['executive_summary']['insights']] +
        [['مشكلة', f'{x["sheet"]}: {x["issue"]} ({x["severity"]})'] for x in report['executive_summary']['issues']] +
        [['توصية', value] for value in report['executive_summary']['recommendations']])
    summary = report['summary']
    dashboard_kpis = [
        ('عدد الأوراق', summary['sheet_count']), ('عدد الصفوف', summary['rows']),
        ('عدد الأعمدة', summary['columns']), ('اكتمال البيانات', f'{summary["completion_rate"]}%'),
        ('الصفوف المكررة', summary['duplicates']), ('درجة الجودة', f'{summary["quality_score"]}/100'),
    ]
    for index, (label, value) in enumerate(dashboard_kpis):
        column = index + 1
        dashboard_sheet.cell(3, column, label); dashboard_sheet.cell(4, column, value)
        dashboard_sheet.cell(3, column).font = Font(bold=True, color='64748B')
        dashboard_sheet.cell(4, column).font = Font(size=18, bold=True, color=BRAND_BLUE)
        dashboard_sheet.cell(3, column).alignment = dashboard_sheet.cell(4, column).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, readingOrder=2)
        dashboard_sheet.cell(3, column).fill = dashboard_sheet.cell(4, column).fill = PatternFill('solid', fgColor='EFF6FF')
        dashboard_sheet.column_dimensions[get_column_letter(column)].width = 19
    dashboard_sheet.row_dimensions[3].height = 25; dashboard_sheet.row_dimensions[4].height = 34
    quality = add_sheet('مؤشرات الجودة', ['المؤشر', 'القيمة'], [[k, v] for k, v in summary.items()])
    add_sheet('معلومات الأوراق', ['الورقة', 'الصفوف', 'الأعمدة', 'الخلايا', 'الاكتمال %', 'الجودة'],
              [[s['name'], s['rows'], s['columns'], s['cells'], s['completion_rate'], s['quality_score']] for s in report['sheets']])
    add_sheet('جودة البيانات', ['الورقة', 'المشكلة', 'الخطورة'], [[s['name'], i['issue'], i['severity']] for s in report['sheets'] for i in s['issues']])
    add_sheet('القيم الفارغة', ['الورقة', 'العمود', 'العدد', 'النسبة %'], [[s['name'], x['column'], x['count'], x['rate']] for s in report['sheets'] for x in s['missing']])
    add_sheet('الصفوف المكررة', ['الورقة', 'العدد', 'النسبة %'], [[s['name'], s['duplicates'], s['duplicate_rate']] for s in report['sheets']])
    add_sheet('أنواع الأعمدة', ['الورقة', 'العمود', 'النوع', 'الفريد', 'ثابت', 'فريد جدًا'], [[s['name'], c['column'], c['type'], c['unique'], c['constant'], c['high_cardinality']] for s in report['sheets'] for c in s['columns_info']])
    stat_keys = ['column','count','sum','mean','median','std','min','q1','q3','max']
    add_sheet('الإحصاءات الرقمية', ['الورقة', *stat_keys], [[s['name'], *[x.get(k, '') for k in stat_keys]] for s in report['sheets'] for x in s['numeric_stats']])
    add_sheet('التحليل النصي', ['الورقة', 'العمود', 'القيم الفريدة', 'أكثر القيم تكرارًا'], [[s['name'], x['column'], x['unique'], '; '.join(f'{v["value"]} ({v["count"]})' for v in x['top_values'])] for s in report['sheets'] for x in s['text_analysis']])
    add_sheet('تحليل التواريخ', ['الورقة', 'العمود', 'العدد', 'الأقدم', 'الأحدث'], [[s['name'], x['column'], x['count'], x['earliest'], x['latest']] for s in report['sheets'] for x in s['date_analysis']])
    add_sheet('القيم الشاذة', ['الورقة', 'العمود', 'العدد', 'النسبة %'], [[s['name'], x['column'], x['count'], x['rate']] for s in report['sheets'] for x in s['outliers']])
    add_sheet('الارتباطات', ['الورقة', 'العمود الأول', 'العمود الثاني', 'الارتباط'], [[s['name'], x['column_1'], x['column_2'], x['value']] for s in report['sheets'] for x in s['correlations']])
    add_sheet('عناصر الداشبورد', ['العنوان', 'النوع', 'الورقة', 'X', 'Y', 'التجميع', 'الألوان'], [[w['title'], w['type'], w['sheet'], w['x_column'], w['y_column'], w['aggregation'], ', '.join(w['colors'])] for w in report['widgets']])
    add_sheet('التوصيات', ['التوصية'], [[x] for x in report['executive_summary']['recommendations']])
    for index, sheet_report in enumerate(report['sheets'], 1):
        frame = sheet_report['preview'].dropna(axis=0, how='all').dropna(axis=1, how='all').head(PREVIEW_ROWS)
        add_sheet(f'معاينة {sheet_report["name"]}', [str(c) for c in frame.columns], frame.values.tolist())

    # Quality chart and internal table of contents.
    info = workbook['معلومات الأوراق']
    if len(report['sheets']):
        chart = BarChart(); chart.title = 'جودة الأوراق'; chart.y_axis.title = 'الدرجة'
        chart.add_data(Reference(info, min_col=6, min_row=1, max_row=info.max_row), titles_from_data=True)
        chart.set_categories(Reference(info, min_col=1, min_row=2, max_row=info.max_row)); quality.add_chart(chart, 'D2')
        dashboard_quality = BarChart(); dashboard_quality.title = 'جودة الأوراق'; dashboard_quality.style = 10
        dashboard_quality.add_data(Reference(info, min_col=6, min_row=1, max_row=info.max_row), titles_from_data=True)
        dashboard_quality.set_categories(Reference(info, min_col=1, min_row=2, max_row=info.max_row))
        dashboard_quality.width = 12; dashboard_quality.height = 7; dashboard_sheet.add_chart(dashboard_quality, 'A7')
    cover.append([]); cover.append(['فهرس التقرير'])
    for title, sheet_name in links[1:]:
        cell = cover.cell(cover.max_row + 1, 1, title); cell.hyperlink = f"#'{sheet_name}'!A1"; cell.style = 'Hyperlink'
    output = io.BytesIO(); workbook.save(output); workbook.close(); output.seek(0); return output


def _encoded_csv(rows, headers):
    text = io.StringIO(newline='')
    writer = csv.writer(
        text,
        delimiter=';',
        quoting=csv.QUOTE_MINIMAL,
        lineterminator='\r\n',
    )
    writer.writerow(headers); writer.writerows(rows)
    return text.getvalue().encode('utf-8-sig')


def _frame_csv(frame):
    text = io.StringIO(newline='')
    frame.to_csv(
        text,
        index=False,
        sep=';',
        quoting=csv.QUOTE_MINIMAL,
        lineterminator='\r\n',
    )
    return text.getvalue().encode('utf-8-sig')


def _safe_filename(value, fallback='sheet'):
    value = re.sub(r'[^\w\-]+', '_', str(value), flags=re.UNICODE).strip('._')
    return value[:80] or fallback


def build_csv_package(dataset, report_data):
    """Export detailed data only: one CSV, or a multi-file ZIP package."""
    report = report_data if 'summary' in report_data else build_report_data(dataset, report_data)
    if len(report['sheets']) == 1:
        frame = report['sheets'][0]['_frame']
        return io.BytesIO(_frame_csv(frame)), 'text/csv; charset=utf-8-sig', '.csv'
    output = io.BytesIO(); used = Counter()
    with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as archive:
        for index, sheet in enumerate(report['sheets'], 1):
            base = _safe_filename(sheet['name']); used[base] += 1
            suffix = f'_{used[base]}' if used[base] > 1 else ''
            archive.writestr(f'{index:02d}_{base}{suffix}.csv', _frame_csv(sheet['_frame']))
        archive.writestr('الملخص_التنفيذي.csv', _encoded_csv([[x] for x in report['executive_summary']['insights']], ['النتيجة']))
        archive.writestr('جودة_البيانات.csv', _encoded_csv([[s['name'], s['quality_score'], s['completion_rate'], s['duplicates']] for s in report['sheets']], ['الورقة','الجودة','الاكتمال','التكرارات']))
        archive.writestr('القيم_الفارغة.csv', _encoded_csv([[s['name'], x['column'], x['count'], x['rate']] for s in report['sheets'] for x in s['missing']], ['الورقة','العمود','العدد','النسبة']))
        archive.writestr('الإحصاءات.csv', _encoded_csv([[s['name'], x['column'], x['count'], x['mean'], x['min'], x['max']] for s in report['sheets'] for x in s['numeric_stats']], ['الورقة','العمود','العدد','المتوسط','الأدنى','الأعلى']))
        archive.writestr('README.txt', 'حزمة Analytix: ملفات البيانات لكل ورقة، والملخص التنفيذي، وجودة البيانات، والقيم الفارغة، والإحصاءات.'.encode('utf-8-sig'))
    output.seek(0); return output, 'application/zip', '.zip'


def prepare_arabic_text(text):
    return get_display(reshape(str(text)))


def get_arabic_font_path():
    fonts_directory = Path(settings.BASE_DIR) / 'static' / 'fonts'
    for filename in ARABIC_FONT_FILENAMES:
        path = fonts_directory / filename
        if path.exists(): return path
    # Matplotlib distributes DejaVu Sans under its permissive license and gives
    # a portable package path on Windows and Render; ReportLab embeds it in PDF.
    bundled = Path(font_manager.findfont('DejaVu Sans', fallback_to_default=False))
    if bundled.exists(): return bundled
    raise ArabicFontNotFoundError('تعذر إنشاء PDF لعدم توفر خط عربي مرخّص. ضع DejaVuSans.ttf في static/fonts/.')


def register_arabic_font():
    font_path = get_arabic_font_path()
    if ARABIC_FONT_NAME not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont(ARABIC_FONT_NAME, str(font_path)))
    return ARABIC_FONT_NAME


def _pdf_table(rows, font, widths=None):
    shaped = [[prepare_arabic_text(v) if isinstance(v, str) else ('' if v is None else v) for v in row] for row in rows]
    table = Table(shaped, colWidths=widths, repeatRows=1, hAlign='RIGHT')
    table.setStyle(TableStyle([
        ('FONTNAME',(0,0),(-1,-1),font), ('ALIGN',(0,0),(-1,-1),'RIGHT'),
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#DBEAFE')), ('TEXTCOLOR',(0,0),(-1,0),colors.HexColor('#1E3A8A')),
        ('GRID',(0,0),(-1,-1),.3,colors.HexColor('#CBD5E1')), ('VALIGN',(0,0),(-1,-1),'TOP'),
        ('FONTSIZE',(0,0),(-1,-1),7), ('LEADING',(0,0),(-1,-1),10),
    ])); return table


def _report_chart(labels, values, title, kind='bar', colors_list=None):
    buffer = io.BytesIO(); figure = None
    try:
        arabic_font_path = get_arabic_font_path()
        font_manager.fontManager.addfont(str(arabic_font_path))
        chart_font = font_manager.FontProperties(fname=str(arabic_font_path))
        plt.rcParams.update({
            'font.size': 10,
            'axes.titlesize': 14, 'axes.labelsize': 10,
            'xtick.labelsize': 9, 'ytick.labelsize': 9,
        })
        figure, axis = plt.subplots(figsize=(9, 4.8), facecolor='#F8FAFC')
        axis.set_facecolor('#FFFFFF')
        palette = colors_list or CHART_COLORS
        display_labels = [prepare_arabic_text(x) for x in labels]
        bar_colors = [palette[i % len(palette)] for i in range(len(values))]
        if kind in ('pie', 'donut'):
            wedges, label_texts, percentage_texts = axis.pie(
                values, labels=display_labels, colors=bar_colors, autopct='%1.1f%%',
                startangle=90, pctdistance=.78,
                wedgeprops={'width': .42 if kind in ('pie', 'donut') else 1, 'edgecolor': 'white'},
            )
            for label in (*label_texts, *percentage_texts):
                label.set_fontproperties(chart_font)
            axis.axis('equal')
        elif kind == 'line':
            axis.plot(display_labels, values, marker='o', linewidth=2.6, markersize=6, color=palette[0])
            axis.fill_between(range(len(values)), values, alpha=.08, color=palette[0])
            axis.grid(axis='y', color='#CBD5E1', linewidth=.7, alpha=.65)
        else:
            positions = range(len(values))
            axis.bar(positions, values, color=bar_colors, width=.62)
            axis.set_xticks(list(positions))
            axis.set_xticklabels(
                display_labels,
                rotation=24 if len(values) > 5 else 0,
                ha='right' if len(values) > 5 else 'center',
            )
            axis.grid(axis='y', color='#CBD5E1', linewidth=.7, alpha=.65)
        if kind not in ('pie', 'donut'):
            axis.set_axisbelow(True)
            axis.yaxis.set_major_formatter(FuncFormatter(lambda value, _: f'{value:,.0f}'))
            for spine in axis.spines.values(): spine.set_visible(False)
            for label in (*axis.get_xticklabels(), *axis.get_yticklabels()):
                label.set_fontproperties(chart_font)
        axis.set_title(
            prepare_arabic_text(title), pad=14, color='#0F172A',
            fontweight='bold', fontproperties=chart_font,
        )
        figure.subplots_adjust(
            left=.10, right=.96,
            bottom=.26 if len(values) > 5 and kind not in ('pie', 'donut') else .14,
            top=.86,
        )
        figure.savefig(buffer, format='png', dpi=300, facecolor=figure.get_facecolor())
        buffer.seek(0)
        return buffer
    except Exception:
        logger.warning('A report chart failed', exc_info=True); buffer.close(); return None
    finally:
        if figure is not None: plt.close(figure)


def _widget_chart(widget, report):
    try:
        sheet = next(item for item in report['sheets'] if item['name'] == widget['sheet'])
        frame = sheet['_frame']
        if widget['x_column'] not in frame.columns or widget['y_column'] not in frame.columns:
            raise KeyError('invalid widget columns')
        grouped = frame.groupby(widget['x_column'], dropna=False)[widget['y_column']]
        if widget['aggregation'] == 'sum': values = grouped.sum()
        elif widget['aggregation'] in ('average', 'mean'): values = grouped.mean()
        else: values = grouped.count()
        values = values.head(TOP_RESULTS)
        return _report_chart(
            [str(x) for x in values.index], values.tolist(),
            widget['title'], 'bar', widget['colors'],
        )
    except Exception:
        report['warnings'].append(f'تعذر إنشاء المخطط: {widget["title"]}')
        logger.warning('Dashboard widget chart failed', exc_info=True)
        return None


def _page_frame(canvas, document):
    canvas.saveState(); canvas.setFont(ARABIC_FONT_NAME, 8); canvas.setFillColor(colors.HexColor('#64748B'))
    canvas.drawRightString(A4[0] - 15*mm, A4[1] - 10*mm, prepare_arabic_text('تقرير Analytix الشامل'))
    canvas.drawCentredString(A4[0] / 2, 9*mm, str(document.page)); canvas.restoreState()


def _pdf_chart_image(chart):
    image = Image(chart, width=165*mm, height=80*mm)
    image.hAlign = 'CENTER'
    return image


def build_pdf_report(dataset, report_data):
    """Build a concise, visual management report for printing and sharing."""
    report = report_data if 'summary' in report_data else build_report_data(dataset, report_data)
    font = register_arabic_font(); output = io.BytesIO(); chart_buffers = []; styles = getSampleStyleSheet()
    body = ParagraphStyle('ArabicBody', parent=styles['BodyText'], fontName=font, alignment=2, fontSize=10.5,
                          leading=16, spaceAfter=7, leftIndent=4, rightIndent=4, allowWidows=0, allowOrphans=0)
    heading = ParagraphStyle('ArabicHeading', parent=body, fontSize=16, leading=22, textColor=colors.HexColor('#1D4ED8'), spaceBefore=10, spaceAfter=7)
    cover = ParagraphStyle('ArabicCover', parent=heading, fontSize=23, leading=30, alignment=1)
    document = SimpleDocTemplate(output, pagesize=A4, rightMargin=14*mm, leftMargin=14*mm, topMargin=18*mm, bottomMargin=16*mm,
                                 title='Analytix comprehensive report')
    p = lambda text, style=body: Paragraph(prepare_arabic_text(text), style)
    story = [Spacer(1, 35*mm), p('Analytix', cover), p('تقرير تحليل البيانات الشامل', cover), Spacer(1, 12*mm),
             _pdf_table([['اسم الملف', report['dataset']['filename']], ['اسم المستخدم', report['user']['username']],
                         ['تاريخ الإنشاء', report['generated_at'].strftime('%Y-%m-%d %H:%M')], ['رقم التقرير', report['dataset']['id']]], font),
             PageBreak(), p('فهرس المحتويات', heading)]
    sections = ['الملخص التنفيذي','مؤشرات التقرير','جودة الأوراق','تفاصيل الأوراق','عناصر الداشبورد','التوصيات النهائية','الملاحق']
    story += [p(f'{i}. {name}') for i, name in enumerate(sections, 1)] + [PageBreak(), p('الملخص التنفيذي', heading)]
    story += [p(x) for x in report['executive_summary']['insights']]
    if report['executive_summary']['issues']:
        story += [p('أهم مشكلات الجودة', heading), _pdf_table([['الورقة','المشكلة','الخطورة'], *[[x['sheet'],x['issue'],x['severity']] for x in report['executive_summary']['issues']]], font)]
    s = report['summary']; story += [p('مؤشرات التقرير', heading), _pdf_table([
        ['الأوراق','الصفوف','الأعمدة','الاكتمال','التكرارات','درجة الجودة'],
        [s['sheet_count'],s['rows'],s['columns'],f'{s["completion_rate"]}%',s['duplicates'],f'{s["quality_score"]}/100']], font)]
    chart = _report_chart([x['name'] for x in report['sheets']], [x['quality_score'] for x in report['sheets']], 'مقارنة جودة الأوراق')
    if chart:
        chart_buffers.append(chart); story += [p('جودة الأوراق', heading), _pdf_chart_image(chart),
            p('المعنى: يقارن الرسم درجة الجودة المحسوبة لكل ورقة.'),
            p(f'الاستنتاج: الجودة العامة للملف {s["quality_score"]} من 100.'),
            p('التوصية: ابدأ بمراجعة الأوراق ذات الدرجة الأقل.')]
    automatic_charts = [
        ('القيم الفارغة', _report_chart([x['name'] for x in report['sheets']], [x['missing_cells'] for x in report['sheets']], 'القيم الفارغة حسب الورقة'),
         f'رُصدت {s["missing_cells"]:,} خلية فارغة.', 'راجع الأوراق والأعمدة الأعلى في القيم الفارغة.'),
        ('أنواع الأعمدة', _report_chart(list(Counter(c['type'] for x in report['sheets'] for c in x['columns_info']).keys()),
                      list(Counter(c['type'] for x in report['sheets'] for c in x['columns_info']).values()), 'توزيع أنواع الأعمدة', 'pie'),
         'يوضح الرسم توزيع أنواع البيانات داخل الملف.', 'تحقق من ملاءمة أنواع الأعمدة لاستخدامها التحليلي.'),
    ]
    numeric_means = [(f'{sheet["name"]}: {stat["column"]}', stat['mean']) for sheet in report['sheets'] for stat in sheet['numeric_stats'] if stat['mean'] is not None][:TOP_RESULTS]
    if numeric_means:
        automatic_charts.append(('أبرز المتوسطات الرقمية', _report_chart([x[0] for x in numeric_means], [x[1] for x in numeric_means], 'أبرز المتوسطات الرقمية'),
                                 'يقارن الرسم متوسطات أهم الأعمدة الرقمية.', 'راجع الفروق الكبيرة في سياق وحدات القياس لكل عمود.'))
    for chart_title, automatic_chart, observation, recommendation in automatic_charts:
        if automatic_chart:
            chart_buffers.append(automatic_chart); story += [p(chart_title, heading), _pdf_chart_image(automatic_chart),
                                                              p(f'الملاحظة: {observation}'), p(f'التوصية: {recommendation}')]
    for sheet in report['sheets']:
        story += [PageBreak(), p(f'الورقة: {sheet["name"]}', heading), _pdf_table([
            ['الصفوف','الأعمدة','الخلايا','الاكتمال','التكرارات','الجودة'],
            [sheet['rows'],sheet['columns'],sheet['cells'],f'{sheet["completion_rate"]}%',sheet['duplicates'],sheet['quality_score']]], font)]
        if sheet['issues']: story += [p('أهم المشاكل', heading), _pdf_table([['المشكلة','الخطورة'], *[[x['issue'],x['severity']] for x in sheet['issues'][:10]]], font)]
        if sheet['missing']: story += [p('القيم الفارغة', heading), _pdf_table([['العمود','العدد','النسبة'], *[[x['column'],x['count'],f'{x["rate"]}%'] for x in sheet['missing'][:10]]], font)]
        if sheet['numeric_stats']: story += [p('الإحصاءات الرقمية', heading), _pdf_table([['العمود','العدد','المتوسط','الوسيط','الأدنى','الأعلى'], *[[x['column'],x['count'],x['mean'],x['median'],x['min'],x['max']] for x in sheet['numeric_stats'][:10]]], font)]
        if sheet['outliers']: story += [p('القيم الشاذة', heading), _pdf_table([['العمود','العدد','النسبة'], *[[x['column'],x['count'],f'{x["rate"]}%'] for x in sheet['outliers'][:10]]], font)]
        story += [p('التوصيات', heading)] + [p(x) for x in sheet['recommendations']]
    if report['widgets']:
        story += [PageBreak(), p('عناصر الداشبورد', heading), _pdf_table([['العنوان','النوع','الورقة','X','Y','التجميع'], *[[w['title'],w['type'],w['sheet'],w['x_column'],w['y_column'],w['aggregation']] for w in report['widgets']]], font)]
        for widget in report['widgets']:
            widget_image = _widget_chart(widget, report)
            if widget_image:
                chart_buffers.append(widget_image); story.append(_pdf_chart_image(widget_image))
    if report['warnings']:
        story += [p('تحذيرات إنشاء التقرير', heading)] + [p(warning) for warning in report['warnings']]
    story += [PageBreak(), p('التوصيات النهائية', heading)] + [p(x) for x in report['executive_summary']['recommendations']]
    story += [p('ملاحظة', heading), p(f'تعرض المعاينات بحد أقصى {PREVIEW_ROWS} صف لكل ورقة، وتبقى البيانات الكاملة محفوظة في ملف المصدر.')]
    try:
        document.build(story, onFirstPage=_page_frame, onLaterPages=_page_frame)
        output.seek(0)
        return output
    finally:
        for chart_buffer in chart_buffers:
            chart_buffer.close()


# Backwards-compatible names for callers outside the datasets app.
build_excel = build_excel_report
build_pdf = build_pdf_report
build_csv = build_csv_package
